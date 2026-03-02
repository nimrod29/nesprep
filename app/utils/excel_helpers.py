"""Excel helper utilities for shift planning."""

import os
from dataclasses import dataclass
from datetime import date
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ShiftSlot:
    """Represents a shift slot in the schedule."""

    day: str
    day_index: int  # 0=Sunday, 6=Saturday
    shift_type: str  # "בוקר", "צהריים", "ערב"
    time_range: str  # e.g., "09:00-18:00"
    row: int
    col: int


@dataclass
class WeekStructure:
    """Structure of a week in the Excel template."""

    start_row: int
    date_range: str
    slots: list[ShiftSlot]


# Hebrew day names
HEBREW_DAYS = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]

# Day column mapping (B=Sunday, H=Saturday)
DAY_COLUMNS = {
    "ראשון": 2,  # B
    "שני": 3,  # C
    "שלישי": 4,  # D
    "רביעי": 5,  # E
    "חמישי": 6,  # F
    "שישי": 7,  # G
    "שבת": 8,  # H
}

# Shift time definitions based on the template
SHIFT_TIMES = {
    "בוקר": ["09:00-18:00"],
    "צהריים": ["10:00-19:00", "11:00-20:00", "12:00-21:00"],
    "ערב": ["13:00-22:00"],
}


def open_workbook(path: str) -> Workbook:
    """Open an Excel workbook."""
    return openpyxl.load_workbook(path)


def save_workbook(wb: Workbook, path: str) -> None:
    """Save an Excel workbook."""
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
    wb.save(path)


def get_sheet_for_month(wb: Workbook, month: int, year: int) -> Worksheet | None:
    """Get the worksheet for a specific month."""
    hebrew_months = {
        1: "ינואר",
        2: "פברואר",
        3: "מרץ",
        4: "אפריל",
        5: "מאי",
        6: "יוני",
        7: "יולי",
        8: "אוגוסט",
        9: "ספטמבר",
        10: "אוקטובר",
        11: "נובמבר",
        12: "דצמבר",
    }

    month_name = hebrew_months.get(month, "")

    for sheet_name in wb.sheetnames:
        if month_name in sheet_name or f"{year}" in sheet_name:
            return wb[sheet_name]

    # Return first sheet if no match
    if wb.sheetnames:
        return wb[wb.sheetnames[0]]
    return None


def find_week_start_row(ws: Worksheet, week_start: date) -> int | None:
    """Find the row where a specific week starts in the worksheet.
    
    Looks for week header patterns like "1-7/3" or "8-14/3" and matches
    based on whether the week_start date falls within that range.
    """
    day = week_start.day
    month = week_start.month

    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and isinstance(cell_value, str):
            # Look for week range pattern like "1-7/3" or "8-14/3"
            if "/" in cell_value and "-" in cell_value:
                try:
                    # Parse "start-end/month" format
                    parts = cell_value.split("/")
                    if len(parts) >= 2:
                        range_part = parts[0]
                        month_part = int(parts[1].split("/")[0]) if "/" in parts[1] else int(parts[1])
                        
                        if "-" in range_part:
                            start_day, end_day = map(int, range_part.split("-"))
                            
                            # Check if the week_start falls within this range
                            if month_part == month and start_day <= day <= end_day:
                                return row
                except (ValueError, IndexError):
                    continue

    return None


def get_week_structure(ws: Worksheet, start_row: int) -> WeekStructure:
    """Parse the structure of a week starting at the given row.
    
    Template structure (offsets from week header row):
    - Row +0: Week label (e.g., "1-7/3")
    - Row +4: Day headers (ראשון, שני, etc.)
    - Row +5: Dates (8.3, 9.3, etc.)
    - Row +7: First shift row (09:00-18:00)
    - Rows +7 to +11: Morning shifts (5 rows)
    - Rows +12 to +14: Afternoon shifts (3 rows)
    - Rows +16 to +21: Evening shifts (6 rows)
    """
    slots: list[ShiftSlot] = []

    # Get date range from the week header row
    date_range = str(ws.cell(row=start_row, column=1).value or "")

    # Morning shifts start at row +7 from week header
    morning_start = start_row + 7
    afternoon_start = start_row + 12
    evening_start = start_row + 16

    # Parse morning slots (09:00-18:00) - 5 rows
    for row_offset in range(5):
        row = morning_start + row_offset
        for day_name, col in DAY_COLUMNS.items():
            slots.append(
                ShiftSlot(
                    day=day_name,
                    day_index=HEBREW_DAYS.index(day_name),
                    shift_type="בוקר",
                    time_range="09:00-18:00",
                    row=row,
                    col=col,
                )
            )

    # Parse afternoon slots (10:00-19:00, 11:00-20:00, 12:00-21:00) - 3 rows
    afternoon_times = ["10:00-19:00", "11:00-20:00", "12:00-21:00"]
    for row_offset in range(3):
        row = afternoon_start + row_offset
        time_range = afternoon_times[row_offset]
        for day_name, col in DAY_COLUMNS.items():
            slots.append(
                ShiftSlot(
                    day=day_name,
                    day_index=HEBREW_DAYS.index(day_name),
                    shift_type="צהריים",
                    time_range=time_range,
                    row=row,
                    col=col,
                )
            )

    # Parse evening slots (13:00-22:00) - 6 rows
    for row_offset in range(6):
        row = evening_start + row_offset
        for day_name, col in DAY_COLUMNS.items():
            slots.append(
                ShiftSlot(
                    day=day_name,
                    day_index=HEBREW_DAYS.index(day_name),
                    shift_type="ערב",
                    time_range="13:00-22:00",
                    row=row,
                    col=col,
                )
            )

    return WeekStructure(start_row=start_row, date_range=date_range, slots=slots)


def get_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Get the value of a cell."""
    return ws.cell(row=row, column=col).value


def set_cell_value(ws: Worksheet, row: int, col: int, value: Any) -> None:
    """Set the value of a cell."""
    ws.cell(row=row, column=col).value = value


def get_assignments_for_week(ws: Worksheet, week_structure: WeekStructure) -> dict[str, list[str]]:
    """Get all current assignments for a week.

    Returns a dict mapping "day|shift_type|time" to list of employee names.
    """
    assignments: dict[str, list[str]] = {}

    for slot in week_structure.slots:
        key = f"{slot.day}|{slot.shift_type}|{slot.time_range}"
        value = get_cell_value(ws, slot.row, slot.col)

        if key not in assignments:
            assignments[key] = []

        if value and isinstance(value, str) and value.strip():
            assignments[key].append(value.strip())

    return assignments


def assign_employee_to_slot(
    ws: Worksheet, slot: ShiftSlot, employee_name: str, append: bool = True
) -> None:
    """Assign an employee to a shift slot.

    If append is True and there's already a value, finds the next empty row in that column.
    """
    current_value = get_cell_value(ws, slot.row, slot.col)

    if not current_value or not append:
        set_cell_value(ws, slot.row, slot.col, employee_name)
    else:
        # Find next empty cell in the same column within the shift type range
        for offset in range(10):  # Check up to 10 rows
            check_row = slot.row + offset
            if not get_cell_value(ws, check_row, slot.col):
                set_cell_value(ws, check_row, slot.col, employee_name)
                return
