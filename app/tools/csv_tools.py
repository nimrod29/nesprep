"""CSV tools for shift planning agents."""

import csv
import io
import json
import os
import shutil
from datetime import date

from langchain_core.tools import tool
from openpyxl import load_workbook

from app.config import settings
from app.utils.csv_template_generator import (
    HEBREW_DAYS,
    generate_month_template,
    get_weeks_in_month,
)

# Shift type to row offsets within a week block (relative to week start)
# Week block starts at row 0 of each week section
SHIFT_ROW_OFFSETS = {
    "בוקר": [6, 7, 8, 9, 10],  # 09:00-18:00 slots
    "צהריים": [11, 12, 13],  # 10:00-19:00, 11:00-20:00, 12:00-21:00
    "ערב": [15, 16, 17, 18, 19, 20],  # 13:00-22:00 slots
}

# Day to column index (0-indexed, column 0 is labels)
DAY_COLUMNS = {
    "ראשון": 1,
    "שני": 2,
    "שלישי": 3,
    "רביעי": 4,
    "חמישי": 5,
    "שישי": 6,
    "שבת": 7,
}

# Row offset for date row within week block
DATE_ROW_OFFSET = 4

# Number of rows per week block (including padding)
WEEK_BLOCK_SIZE = 27

# Header rows before first week block
HEADER_ROWS = 6


class CSVTools:
    """Tools for CSV-based shift planning."""

    def __init__(
        self,
        shift_plan_id: int,
        output_dir: str | None = None,
    ):
        self.shift_plan_id = shift_plan_id
        self.output_dir = output_dir or settings.OUTPUT_DIR
        self._csv_data: list[list[str]] | None = None
        self._csv_path: str | None = None
        self._year: int | None = None
        self._month: int | None = None

    def get_tools(self) -> list:
        """Return the list of CSV tools."""
        tools_self = self

        @tool
        def get_current_date() -> str:
            """Get today's date and day of week for planning context.

            Returns:
                JSON with date, day_name (Hebrew), and day_index (0=Sunday).
            """
            today = date.today()
            # Python weekday: Monday=0, Sunday=6
            # Convert to Sunday=0, Saturday=6
            day_index = (today.weekday() + 1) % 7
            day_name = HEBREW_DAYS[day_index]

            return json.dumps(
                {
                    "date": today.isoformat(),
                    "day": today.day,
                    "month": today.month,
                    "year": today.year,
                    "day_name": day_name,
                    "day_index": day_index,
                },
                ensure_ascii=False,
            )

        @tool
        def generate_template(year: int, month: int) -> str:
            """Generate a fresh CSV template for the specified month.

            This creates a new template with correct dates for all weeks
            that overlap with the given month.

            Args:
                year: The year (e.g., 2025)
                month: The month (1-12)

            Returns:
                JSON with path to generated file and summary.
            """
            try:
                csv_content = generate_month_template(year, month)
                tools_self._csv_data = list(csv.reader(io.StringIO(csv_content)))
                tools_self._year = year
                tools_self._month = month

                os.makedirs(tools_self.output_dir, exist_ok=True)
                filename = f"shift_plan_{year}_{month:02d}.csv"
                tools_self._csv_path = os.path.join(tools_self.output_dir, filename)

                with open(tools_self._csv_path, "w", encoding="utf-8", newline="") as f:
                    f.write(csv_content)

                weeks = get_weeks_in_month(year, month)
                week_ranges = [f"{w[0].day}.{w[0].month}-{w[1].day}.{w[1].month}" for w in weeks]

                return json.dumps(
                    {
                        "success": True,
                        "path": tools_self._csv_path,
                        "year": year,
                        "month": month,
                        "weeks": len(weeks),
                        "week_ranges": week_ranges,
                        "total_rows": len(tools_self._csv_data),
                    },
                    ensure_ascii=False,
                )
            except Exception as e:
                return json.dumps({"success": False, "error": str(e)})

        @tool
        def read_csv() -> str:
            """Read the current CSV template as formatted text.

            Shows the template structure with row numbers for reference.
            Use this to understand the current state before making assignments.

            Returns:
                Formatted text representation of the CSV.
            """
            if tools_self._csv_data is None:
                return "Error: No template loaded. Call generate_template first."

            lines = []
            for i, row in enumerate(tools_self._csv_data):
                row_str = ",".join(row)
                lines.append(f"{i:3d}: {row_str}")

            return "\n".join(lines)

        @tool
        def assign_shift(employee_name: str, day_date: str, shift_type: str) -> str:
            """Assign an employee to a shift on a specific date.

            Args:
                employee_name: Name of the employee (in Hebrew)
                day_date: Date as "D.M" format (e.g., "2.3" for March 2nd)
                shift_type: One of בוקר (morning), צהריים (afternoon), ערב (evening)

            Returns:
                Confirmation message or error.
            """
            if tools_self._csv_data is None:
                return "Error: No template loaded. Call generate_template first."

            if shift_type not in SHIFT_ROW_OFFSETS:
                return f"Error: Invalid shift type '{shift_type}'. Must be בוקר, צהריים, or ערב"

            # Find the week block and column for this date
            week_block_start = None
            col_index = None

            for row_idx, row in enumerate(tools_self._csv_data):
                if len(row) > 0 and row[0] == "תאריך":
                    # This is a date row - check if our date is in it
                    for col_idx, cell in enumerate(row[1:], start=1):
                        if cell == day_date:
                            # Found the date - calculate week block start
                            # Date row is at offset 4 within week block
                            week_block_start = row_idx - DATE_ROW_OFFSET
                            col_index = col_idx
                            break
                    if week_block_start is not None:
                        break

            if week_block_start is None or col_index is None:
                return f"Error: Date '{day_date}' not found in template"

            # Find an empty slot for this shift type
            row_offsets = SHIFT_ROW_OFFSETS[shift_type]
            assigned = False

            for offset in row_offsets:
                row_idx = week_block_start + offset
                if row_idx < len(tools_self._csv_data):
                    row = tools_self._csv_data[row_idx]
                    # Ensure row has enough columns
                    while len(row) <= col_index:
                        row.append("")

                    if row[col_index] == "":
                        row[col_index] = employee_name
                        assigned = True
                        break

            if not assigned:
                return f"Error: No empty slot available for {shift_type} on {day_date}"

            return f"Assigned {employee_name} to {shift_type} on {day_date}"

        @tool
        def get_assignments_summary() -> str:
            """Get a summary of all current assignments in the template.

            Returns:
                JSON with assignments grouped by week and day.
            """
            if tools_self._csv_data is None:
                return "Error: No template loaded. Call generate_template first."

            assignments: dict[str, dict[str, list[str]]] = {}

            # Find all date rows and collect assignments
            for row_idx, row in enumerate(tools_self._csv_data):
                if len(row) > 0 and row[0] == "תאריך":
                    week_block_start = row_idx - DATE_ROW_OFFSET
                    dates = row[1:8]  # Get dates for this week

                    for col_idx, date_str in enumerate(dates, start=1):
                        if not date_str:
                            continue

                        if date_str not in assignments:
                            assignments[date_str] = {"בוקר": [], "צהריים": [], "ערב": []}

                        # Check each shift type
                        for shift_type, offsets in SHIFT_ROW_OFFSETS.items():
                            for offset in offsets:
                                check_row = week_block_start + offset
                                if check_row < len(tools_self._csv_data):
                                    check_row_data = tools_self._csv_data[check_row]
                                    if col_idx < len(check_row_data) and check_row_data[col_idx]:
                                        assignments[date_str][shift_type].append(
                                            check_row_data[col_idx]
                                        )

            # Calculate totals
            total_assignments = sum(
                len(shifts)
                for day_data in assignments.values()
                for shifts in day_data.values()
            )

            return json.dumps(
                {
                    "assignments": assignments,
                    "total_assignments": total_assignments,
                    "days_with_assignments": len(
                        [d for d in assignments.values() if any(s for s in d.values())]
                    ),
                },
                ensure_ascii=False,
                indent=2,
            )

        @tool
        def save_csv() -> str:
            """Save the current state to the CSV file.

            Returns:
                Path to the saved file or error message.
            """
            if tools_self._csv_data is None or tools_self._csv_path is None:
                return "Error: No template loaded. Call generate_template first."

            try:
                with open(tools_self._csv_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerows(tools_self._csv_data)

                return f"Saved CSV to: {tools_self._csv_path}"
            except Exception as e:
                return f"Error saving CSV: {e}"

        @tool
        def convert_to_excel() -> str:
            """Convert the CSV data to a formatted Excel file using the template.

            Copies the original Excel template and fills in:
            - Week date range titles
            - Dates for each week
            - Employee assignments in the correct cells

            Returns:
                Path to the Excel file or error message.
            """
            if tools_self._csv_data is None:
                return "Error: No template loaded. Call generate_template first."

            try:
                # Get template path
                template_path = settings.get_template_path()
                if not os.path.exists(template_path):
                    return f"Error: Template not found at {template_path}"

                # Copy template to output
                excel_filename = f"shift_plan_{tools_self._year}_{tools_self._month:02d}.xlsx"
                excel_path = os.path.join(tools_self.output_dir, excel_filename)
                shutil.copy2(template_path, excel_path)

                # Open the copied template
                wb = load_workbook(excel_path)
                ws = wb.active
                if ws is None:
                    return "Error: Could not open worksheet"

                # Get weeks for this month
                weeks = get_weeks_in_month(tools_self._year, tools_self._month)

                # Template structure: each week block starts at row 6 (title), row 7 (events), etc.
                # Week title is at row 6, 33, 60, etc. (27 rows per block, starting at 6)
                # Row offsets within block (from title row):
                # - Row 0: Week date range title (e.g., "1-7/2/2026")
                # - Row 1: אירועי השבוע
                # - Row 5: Date row (תאריך)
                # - Rows 7-11: Morning shifts
                # - Rows 12-14: Afternoon shifts
                # - Rows 16-21: Evening shifts

                TEMPLATE_WEEK_TITLE_START = 6  # First week title at row 6
                TEMPLATE_WEEK_SIZE = 27  # Each week block is 27 rows
                DATE_ROW_OFFSET_IN_BLOCK = 5  # Date row is 5 rows after title

                # Map CSV data to Excel
                csv_week_idx = 0
                for row_idx, row in enumerate(tools_self._csv_data):
                    if len(row) > 0 and row[0] == "תאריך":
                        # This is a date row in CSV - find corresponding Excel row
                        excel_week_title_row = TEMPLATE_WEEK_TITLE_START + (csv_week_idx * TEMPLATE_WEEK_SIZE)
                        excel_date_row = excel_week_title_row + DATE_ROW_OFFSET_IN_BLOCK

                        # Get the week start and end dates from CSV row
                        dates_in_week = [d for d in row[1:8] if d]
                        if dates_in_week:
                            first_date = dates_in_week[0]  # e.g., "29.3"
                            last_date = dates_in_week[-1]  # e.g., "4.4"
                            
                            # Parse dates to create range title
                            first_parts = first_date.split(".")
                            last_parts = last_date.split(".")
                            
                            if first_parts[1] == last_parts[1]:
                                # Same month: "29-4/4/2026"
                                week_title = f"{first_parts[0]}-{last_parts[0]}/{last_parts[1]}/{tools_self._year}"
                            else:
                                # Different months: "29/3-4/4/2026"
                                week_title = f"{first_parts[0]}/{first_parts[1]}-{last_parts[0]}/{last_parts[1]}/{tools_self._year}"
                            
                            # Update week title
                            ws.cell(row=excel_week_title_row, column=1, value=week_title)

                        # Fill in dates (columns B-H, which is 2-8)
                        for col_idx in range(1, 8):
                            if col_idx < len(row):
                                ws.cell(row=excel_date_row, column=col_idx + 1, value=row[col_idx])

                        # Fill in assignments from the CSV data
                        csv_week_start = row_idx - DATE_ROW_OFFSET
                        excel_week_start = excel_week_title_row + 1  # Events row is 1 after title

                        for shift_type, offsets in SHIFT_ROW_OFFSETS.items():
                            for offset in offsets:
                                csv_row_idx = csv_week_start + offset
                                excel_row = excel_week_start + offset

                                if csv_row_idx < len(tools_self._csv_data):
                                    csv_row = tools_self._csv_data[csv_row_idx]
                                    for col_idx in range(1, 8):
                                        if col_idx < len(csv_row) and csv_row[col_idx]:
                                            ws.cell(row=excel_row, column=col_idx + 1, value=csv_row[col_idx])

                        csv_week_idx += 1

                # Save
                wb.save(excel_path)

                return f"Converted to Excel: {excel_path}"
            except Exception as e:
                return f"Error converting to Excel: {e}"

        return [
            get_current_date,
            generate_template,
            read_csv,
            assign_shift,
            get_assignments_summary,
            save_csv,
            convert_to_excel,
        ]
